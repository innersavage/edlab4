from sklearn import cluster
import pandas
import matplotlib.pyplot as plt
import matplotlib.cm
import matplotlib.markers
import matplotlib.lines
import matplotlib.colors
from sklearn.decomposition import PCA
import seaborn
import numpy as np
import time
import openpyxl
from sklearn.metrics import auc, roc_curve, roc_auc_score
import matplotlib.pyplot as plt
import matplotlib.markers as markers
import matplotlib.colors as colors

from sklearn import svm
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier, VotingClassifier

from sklearn.metrics import confusion_matrix
from sklearn.metrics import accuracy_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
from sklearn.metrics import roc_auc_score

from sklearn.model_selection import cross_val_score

# Zadanie 1 (1 pkt.)
print('\n## W pocie czoła ładuję dane...')
X_train = pandas.read_csv("X_train.txt", sep=' ', header=None)
X_test = pandas.read_csv("X_test.txt", sep=' ', header=None)
y_train = pandas.read_csv("y_train.txt", sep=' ', header=None)
y_test = pandas.read_csv("y_test.txt", sep=' ', header=None)


# Zadanie 2 (2 pkt.)
pca_10d = PCA(n_components=10)
X_train_decomposed = pca_10d.fit_transform(X_train)
X_test_decomposed = pca_10d.fit_transform(X_test)
t0 = time.process_time_ns()
clf = svm.SVC(kernel='linear').fit(X_train, y_train.values.ravel())
t1 = time.process_time_ns()
print('Klasyfikator SVM przed redukcją wymiarowości, czas trenowania (w nanosekundach): {}'.format(t1 - t0))
t0 = time.process_time_ns()
clf_dec = svm.SVC(kernel='linear').fit(X_train_decomposed, y_train.values.ravel())
t1 = time.process_time_ns()
print('Klasyfikator SVM po redukcji wymiarowości, czas trenowania (w nanosekundach): {}'.format(t1 - t0))
t0 = time.process_time_ns()
scores1 = cross_val_score(clf, X_test, y_test.values.ravel(), cv=5, n_jobs=-1)
t1 = time.process_time_ns()
print('Klasyfikator SVM przed redukcją wymiarowości, czas testowania cross_val_score (w nanosekundach): {}'.format(t1 - t0))
t0 = time.process_time_ns()
scores2 = cross_val_score(clf_dec, X_test_decomposed, y_test.values.ravel(), cv=5, n_jobs=-1)
t1 = time.process_time_ns()
print('Klasyfikator SVM po redukcji wymiarowości, czas testowania cross_val_score (w nanosekundach): {}'.format(t1 - t0))
t0 = time.process_time_ns()
y_pred = clf.predict(X_test)
scores3 = accuracy_score(y_test.values.ravel(), y_pred)
t1 = time.process_time_ns()
print('Klasyfikator SVM przed redukcją wymiarowości, czas testowania ACC (w nanosekundach): {}'.format(t1 - t0))
t0 = time.process_time_ns()
y_pred = clf_dec.predict(X_test_decomposed)
scores4 = accuracy_score(y_test.values.ravel(), y_pred)
t1 = time.process_time_ns()
print('Klasyfikator SVM po redukcji wymiarowości, czas testowania ACC (w nanosekundach): {}'.format(t1 - t0))

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Klasyfikator SVM przed redukcją wymiarowości cross_val_score'
ws['B1'] = 'Klasyfikator SVM po redukcji wymiarowości cross_val_score'
ws['C1'] = 'Klasyfikator SVM przed redukcją wymiarowości ACC'
ws['D1'] = 'Klasyfikator SVM po redukcji wymiarowości ACC'
for i, j in [(scores1, 'A'), (scores2, 'B'), ([scores3], 'C'), ([scores4], 'D')]:
    for k, l in enumerate(i, start=2):
        ws[j + str(k)] = l
wb.save('dim_reduction.xlsx')


# Zadanie 3 (1 pkt.)
model_svm = svm.SVC(probability=True)
model_svm.fit(X_train, y_train.values.ravel())
model_knn = KNeighborsClassifier()
model_knn.fit(X_train, y_train.values.ravel())
model_decision_tree = DecisionTreeClassifier()
model_decision_tree.fit(X_train, y_train.values.ravel())
model_random_forest = RandomForestClassifier()
model_random_forest.fit(X_train, y_train.values.ravel())

estimators = [('SVM', model_svm),
              ('KNN', model_knn),
              ('Decision Tree', model_decision_tree),
              ('Random Forest', model_random_forest)]
ensemble = VotingClassifier(estimators=estimators, voting='soft', weights=[0.3, 0.3, 0.2, 0.2])
ensemble.fit(X_train_decomposed, y_train.values.ravel())

print("Wyniki dla zespołu klasyfikacyjnego")
encoded_features = pandas.get_dummies(y_test.values.ravel())
y_pred = ensemble.predict(X_test_decomposed)
acc = accuracy_score(y_test.values.ravel(), y_pred)
recall = recall_score(y_test.values.ravel(), y_pred, average='macro')
f1 = f1_score(y_test.values.ravel(), y_pred, average='macro')

fpr, tpr, thresh = {}, {}, {}
n_class = len(np.unique(y_train))
y_pred_proba = ensemble.predict_proba(X_test_decomposed)
auc_scores = []
for i in range(n_class):
        fpr[i], tpr[i], thresh[i] = roc_curve(y_test.values.ravel(), y_pred_proba[:,i], pos_label=i)
        auc_score = roc_auc_score(encoded_features.iloc[:,i], y_pred_proba[:,i],multi_class='ovr', average='macro')
        auc_scores.append(auc_score)

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'ACC'
ws['B1'] = 'Recall'
ws['C1'] = 'F1'
ws['D1'] = 'AUC'
ws['A2'] = acc
ws['B2'] = recall
ws['C2'] = f1
for i, j in enumerate(auc_scores, start=2):
    ws['C' + str(i)] = j
wb.save('ensambled_learning.xlsx')

# Zadanie 4
train_dataset = X_train
train_dataset['Class'] = pandas.Series(y_train.values.ravel())
test_dataset = X_test
test_dataset['Class'] = pandas.Series(y_test.values.ravel())
mrks = list(markers.MarkerStyle.markers.keys())[0:n_class]
clrs = list(colors.XKCD_COLORS.keys())
plot_step = 0.05
ensemble.fit(X_train_decomposed[:,:2],y_train.values.ravel())

train_predicted = ensemble.predict(X_train_decomposed[:,:2])
train_shrinked_df = pandas.DataFrame(X_train_decomposed)
train_shrinked_df['Class'] = pandas.Series(train_predicted.tolist())

test_predicted = ensemble.predict(X_test_decomposed[:,:2])
test_shrinked_df = pandas.DataFrame(X_test_decomposed)
test_shrinked_df['Class'] = pandas.Series(test_predicted.tolist())

x_min, x_max = float(train_shrinked_df.iloc[:, 0].min()) - 1, float(train_shrinked_df.iloc[:, 0].max()) + 1
y_min, y_max = float(train_shrinked_df.iloc[:, 1].min()) - 1, float(train_shrinked_df.iloc[:, 1].max()) + 1
xx, yy = np.meshgrid(np.arange(x_min, x_max, plot_step), \
                        np.arange(y_min, y_max, plot_step))
Z = ensemble.predict(np.c_[xx.ravel(), yy.ravel()])
Z = Z.reshape(xx.shape)
fig, ax = plt.subplots(nrows=3,ncols=1, sharex=False, figsize=(15,15))
ax[1].contourf(xx, yy, Z)
x_min, x_max = float(test_shrinked_df.iloc[:, 0].min()) - 1, float(test_shrinked_df.iloc[:, 0].max()) + 1
y_min, y_max = float(test_shrinked_df.iloc[:, 1].min()) - 1, float(test_shrinked_df.iloc[:, 1].max()) + 1
print(x_min, x_max)
print(y_min, y_max)
xx, yy = np.meshgrid(np.arange(x_min, x_max, plot_step),\
                         np.arange(y_min, y_max, plot_step))
Z = ensemble.predict(np.c_[xx.ravel(), yy.ravel()])
Z = Z.reshape(xx.shape)
ax[2].contourf(xx, yy, Z)
for i in range(n_class):
    sliced_train = train_dataset.loc[train_dataset.Class == i]
    sliced_test = test_dataset.loc[test_dataset.Class == i]
    sliced_train.plot(x=0,y=1, ax=ax[0], kind='scatter', marker=mrks[i-1], c=clrs[i-1], label="Klasa "+str(i))
    sliced_test.plot(x=0,y=1, ax=ax[0], kind='scatter', marker=mrks[i-1], c=clrs[i-1], label="Test-Klasa"+str(i),edgecolor='black', linewidth=1)
ax[0].legend(bbox_to_anchor=(1.2, 1.3))
ax[0].set_title("Rozkład próbek treningowych / testowych z podziałem na klasy")
ax[1].set_title("Wynik trenowania modelu z podziałem na klasy")
ax[2].set_title("Wynik testowania modelu z podziałem na klasy")
for i in range(n_class):
    current = train_shrinked_df.loc[train_shrinked_df.Class == i]
    current.plot(x=0,y=2, ax=ax[1], kind='scatter', marker=mrks[i-1], c=clrs[i-1], label="Klasa"+str(i), zorder=1)
    current = test_shrinked_df.loc[test_shrinked_df.Class == i]
    current.plot(x=0,y=2, ax=ax[2], kind='scatter', marker=mrks[i-1], c=clrs[i-1], label="Klasa"+str(i), zorder=1)
plt.show()